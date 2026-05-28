"""
Zorlu Enerji – Plant-Level Monthly Electricity Generation
=========================================================
Source  : EPİAŞ Transparency Platform  (seffaflik.epias.com.tr)
Endpoint: rt-gen-bulk  (Santral Bazlı Toplu Gerçek Zamanlı Üretim)
Coverage: May 2019 – present  (plant-level data published from 16 May 2019)
Output  : zorlu_enerji_generation.xlsx  +  intermediate CSVs
"""

import os, time
from datetime import date, timedelta
from pathlib import Path

import pandas as pd
from tqdm import tqdm
from dotenv import load_dotenv
from eptr2 import EPTR2

load_dotenv()

# ── Output paths ──────────────────────────────────────────────────────────────
OUT_DIR   = Path(__file__).parent
RAW_CSV   = OUT_DIR / "zorlu_generation_raw.csv"
EXCEL_OUT = OUT_DIR / "zorlu_enerji_generation.xlsx"

# ── Confirmed Zorlu Enerji physical plant IDs (from EPIAS pp-list) ────────────
# Verified by cross-referencing with 2024 Annual Report plant list.
# IDs are the physical plant (santral) identifiers used in rt-gen-bulk.
ZORLU_PLANTS = [
    {"pp_id": 1400, "plant_name": "Kızıldere I",      "fuel_type": "Geothermal",    "capacity_mw": 15.00,  "location": "Denizli"},
    {"pp_id": 1628, "plant_name": "Kızıldere II",     "fuel_type": "Geothermal",    "capacity_mw": 80.00,  "location": "Denizli"},
    {"pp_id": 2215, "plant_name": "Kızıldere III",    "fuel_type": "Geothermal",    "capacity_mw": 165.00, "location": "Denizli-Aydın"},
    {"pp_id": 1926, "plant_name": "Alaşehir",         "fuel_type": "Geothermal",    "capacity_mw": 45.00,  "location": "Manisa"},
    {"pp_id": 2283, "plant_name": "İkizdere",         "fuel_type": "Hydroelectric", "capacity_mw": 24.94,  "location": "Rize"},
    {"pp_id": 1353, "plant_name": "Tercan",           "fuel_type": "Hydroelectric", "capacity_mw": 15.00,  "location": "Erzincan"},
    {"pp_id": 1002, "plant_name": "Mercan",           "fuel_type": "Hydroelectric", "capacity_mw": 20.40,  "location": "Tunceli"},
    {"pp_id": 2212, "plant_name": "Mercan (Yukarı)",  "fuel_type": "Hydroelectric", "capacity_mw": None,   "location": "Tunceli"},
    {"pp_id": 2678, "plant_name": "Mercan (Hacı)",    "fuel_type": "Hydroelectric", "capacity_mw": None,   "location": "Tunceli"},
    {"pp_id": 943,  "plant_name": "Beyköy",           "fuel_type": "Hydroelectric", "capacity_mw": 16.80,  "location": "Eskişehir"},
    {"pp_id": 1340, "plant_name": "Kuzgun",           "fuel_type": "Hydroelectric", "capacity_mw": 20.90,  "location": "Erzurum"},
    {"pp_id": 1287, "plant_name": "Çıldır",           "fuel_type": "Hydroelectric", "capacity_mw": 15.40,  "location": "Kars"},
    {"pp_id": 1370, "plant_name": "Ataköy",           "fuel_type": "Hydroelectric", "capacity_mw": 5.50,   "location": "Tokat"},
    {"pp_id": 1177, "plant_name": "Gökçedağ",        "fuel_type": "Wind",          "capacity_mw": 135.00, "location": "Osmaniye"},
    {"pp_id": 1222, "plant_name": "Lüleburgaz",       "fuel_type": "Natural Gas",   "capacity_mw": 49.53,  "location": "Kırklareli"},
]

PP_IDS = [p["pp_id"] for p in ZORLU_PLANTS]

# Gökçedağ (pp_id 1177) was sold to Rönesans Enerji in December 2025.
# Production data after this date belongs to the new owner, not Zorlu.
GOKCEDAG_PP_ID   = 1177
GOKCEDAG_SALE_DT = date(2025, 12, 31)   # last day to include for Zorlu

# Plant-level data available from this date
DATA_START = date(2019, 5, 16)


# ── EPIAS client ──────────────────────────────────────────────────────────────

def get_client() -> EPTR2:
    u = os.getenv("EPTR_USERNAME")
    p = os.getenv("EPTR_PASSWORD")
    if not u or not p:
        raise ValueError("Fill EPTR_USERNAME and EPTR_PASSWORD in .env")
    return EPTR2(username=u, password=p)


# ── Data fetching ─────────────────────────────────────────────────────────────

def fetch_day(eptr: EPTR2, day: date, retries: int = 3) -> pd.DataFrame:
    """
    Fetch hourly generation for all Zorlu plants for one calendar day.
    Returns a DataFrame with columns: date, hour, pp_id, plant_name, total_mwh.
    Returns empty DataFrame on failure.
    """
    date_str = f"{day.isoformat()}T00:00:00+03:00"
    for attempt in range(retries):
        try:
            df = eptr.call("rt-gen-bulk", date=date_str, pp_ids=PP_IDS)
            if df is None or df.empty:
                return pd.DataFrame()

            # Extract plant id from the powerPlantName column (format: "NAME-EIC-ID")
            if "powerPlantName" in df.columns:
                df["pp_id"] = df["powerPlantName"].str.extract(r"-(\d+)$").astype(int)
            else:
                return pd.DataFrame()

            keep = df[df["pp_id"].isin(PP_IDS)][["date", "hour", "pp_id", "total"]].copy()
            keep.rename(columns={"total": "production_mwh"}, inplace=True)
            keep["production_mwh"] = pd.to_numeric(keep["production_mwh"], errors="coerce").fillna(0)
            return keep

        except Exception:
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
    return pd.DataFrame()


def build_raw_data(eptr: EPTR2) -> pd.DataFrame:
    """
    Loop day by day from DATA_START to yesterday.
    Saves a checkpoint CSV every 50 days so the run can be resumed.
    """
    yesterday = date.today() - timedelta(days=1)
    all_days  = [DATA_START + timedelta(n) for n in range((yesterday - DATA_START).days + 1)]

    # Resume: find the last date already saved
    done_dates: set[date] = set()
    saved_rows: list[dict] = []
    if RAW_CSV.exists():
        existing = pd.read_csv(RAW_CSV, parse_dates=["date"])
        done_dates = set(existing["date"].dt.date)
        saved_rows = existing.to_dict("records")
        print(f"  Resuming — {len(done_dates)} days already fetched.")

    remaining = [d for d in all_days if d not in done_dates]
    if not remaining:
        print("  All days already fetched.")
        return pd.read_csv(RAW_CSV)

    new_rows: list[dict] = []
    with tqdm(total=len(remaining), desc="  Fetching", unit="day") as pbar:
        for i, day in enumerate(remaining):
            df_day = fetch_day(eptr, day)
            if not df_day.empty:
                new_rows.extend(df_day.to_dict("records"))
            pbar.update(1)
            time.sleep(0.35)

            # Checkpoint every 50 days
            if (i + 1) % 50 == 0:
                pd.DataFrame(saved_rows + new_rows).to_csv(RAW_CSV, index=False)

    all_rows = saved_rows + new_rows
    df = pd.DataFrame(all_rows)
    df.to_csv(RAW_CSV, index=False)
    print(f"  OK: Raw data: {len(df)} rows saved -> {RAW_CSV.name}")
    return df


# ── Aggregation ───────────────────────────────────────────────────────────────

def aggregate_to_monthly(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Sum hourly MWh to monthly MWh per physical plant.

    Gökçedağ rows after GOKCEDAG_SALE_DT are excluded: that plant was sold
    in December 2025 and subsequent output belongs to the new owner.
    """
    df = df_raw.copy()
    df["date"] = pd.to_datetime(df["date"])
    df["year"]  = df["date"].dt.year
    df["month"] = df["date"].dt.month
    df["pp_id"] = df["pp_id"].astype(int)

    # Drop post-sale Gökçedağ rows
    mask_gokcedag_postsale = (
        (df["pp_id"] == GOKCEDAG_PP_ID) &
        (df["date"].dt.date > GOKCEDAG_SALE_DT)
    )
    df = df[~mask_gokcedag_postsale]

    monthly = (
        df.groupby(["year", "month", "pp_id"], as_index=False)["production_mwh"]
        .sum()
        .round(1)
    )

    # Attach plant metadata
    meta_df = pd.DataFrame(ZORLU_PLANTS)
    monthly = monthly.merge(meta_df, on="pp_id", how="left")
    # Period as "YYYY-MM" string — clean for monthly data, no time component
    monthly["period"] = monthly["year"].astype(str) + "-" + monthly["month"].astype(str).str.zfill(2)
    return monthly


# ── Excel export ──────────────────────────────────────────────────────────────

def export_excel(monthly: pd.DataFrame) -> None:
    print("\n[4/4] Building Excel ...")
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    pull_date = pd.Timestamp.now().strftime("%Y-%m-%d")

    with pd.ExcelWriter(EXCEL_OUT, engine="openpyxl") as writer:

        # ── Tab 0: Guide — table of contents explaining each sheet ───────
        guide_data = [
            ("Sheet",                "What it contains",
             "Best used for"),
            ("Summary",             "One row per month — total production of ALL Zorlu plants combined (MWh)",
             "Quick top-line view of the company. Spot overall trends and seasonality."),
            ("By Plant",            "One row per plant per month — the main data table. Includes plant name, fuel type, installed capacity (MW), location and production (MWh)",
             "Primary analysis sheet. Filter by plant or fuel type, calculate capacity factors, build any custom view."),
            ("Wide Pivot",          "Same data as By Plant, but reshaped: rows = months, columns = individual plants",
             "Plug directly into an Excel chart to compare plants visually side by side."),
            ("By Fuel Type",        "Monthly production grouped by energy source: Geothermal / Hydroelectric / Wind / Natural Gas",
             "Track how the generation mix has shifted over time (e.g. gas phased out after 2020)."),
            ("Raw Hourly (sample)", "First 50,000 hourly readings straight from the EPIAS API, before any aggregation",
             "Audit or verify the source data. Not needed for normal analysis."),
            ("Metadata",            "Data source, API endpoint, date pulled, coverage dates and known caveats",
             "Read before using the data. Lists what is NOT included and why."),
            ("Plant Reference",     "Static list of all 15 plants with EPIAS ID, fuel type, installed capacity (MW) and location",
             "Quick lookup of plant characteristics or to reproduce an API query for a specific plant."),
        ]
        guide_df = pd.DataFrame(guide_data[1:], columns=list(guide_data[0]))
        guide_df.to_excel(writer, sheet_name="Guide", index=False)

        # ── Tab 1: Summary — one row per month, all plants combined ──────
        # Useful as a quick top-line view: total company output each month.
        summary = (
            monthly.groupby(["period", "year", "month"], as_index=False)["production_mwh"]
            .sum()
            .rename(columns={"period": "Year-Month",
                             "production_mwh": "Total Production (MWh)"})
            .sort_values(["year", "month"])
        )
        summary[["Year-Month", "Total Production (MWh)"]].to_excel(
            writer, sheet_name="Summary", index=False
        )

        # ── Tab 2: By Plant — main analyst tab ───────────────────────────
        # One row per (month × plant). Contains all key columns for analysis:
        # Plant Name, Fuel Type, Installed Capacity, and monthly production.
        long = (
            monthly.rename(columns={
                "period":        "Year-Month",
                "plant_name":    "Plant Name",
                "fuel_type":     "Fuel Type",
                "capacity_mw":   "Installed Capacity (MW)",
                "location":      "Location",
                "production_mwh": "Production (MWh)",
            })
            .sort_values(["year", "month", "Plant Name"])
        )
        long[["Year-Month", "Plant Name", "Fuel Type",
              "Installed Capacity (MW)", "Location", "Production (MWh)"]].to_excel(
            writer, sheet_name="By Plant", index=False
        )

        # ── Tab 3: Wide Pivot — months as rows, plants as columns ─────────
        # One row per month, one column per plant. Easy for charting in Excel
        # and for quick comparison across plants.
        wide = monthly.pivot_table(
            index=["period", "year", "month"],
            columns="plant_name",
            values="production_mwh",
            aggfunc="sum",
            fill_value=0,
        ).reset_index()
        wide.columns.name = None
        wide = wide.rename(columns={"period": "Year-Month"})
        wide = wide.sort_values(["year", "month"]).drop(columns=["year", "month"])
        wide.to_excel(writer, sheet_name="Wide Pivot", index=False)

        # ── Tab 4: By Fuel Type — monthly totals grouped by energy source ─
        # Geothermal / Hydroelectric / Wind / Natural Gas broken out separately.
        by_fuel = (
            monthly.groupby(["period", "year", "month", "fuel_type"], as_index=False)
            ["production_mwh"].sum()
            .rename(columns={"period": "Year-Month",
                             "fuel_type": "Fuel Type",
                             "production_mwh": "Production (MWh)"})
            .sort_values(["year", "month", "Fuel Type"])
        )
        by_fuel[["Year-Month", "Fuel Type", "Production (MWh)"]].to_excel(
            writer, sheet_name="By Fuel Type", index=False
        )

        # ── Tab 5: Raw hourly sample — first 50k rows of source data ──────
        # The actual hourly readings from EPIAS before monthly aggregation.
        # pp_id is the EPIAS physical plant ID (see Plant Reference tab for mapping).
        raw = pd.read_csv(RAW_CSV)
        raw["date"] = pd.to_datetime(raw["date"]).dt.strftime("%Y-%m-%d")
        # Add plant name for readability
        id_to_name = {p["pp_id"]: p["plant_name"] for p in ZORLU_PLANTS}
        raw["Plant Name"] = raw["pp_id"].map(id_to_name)
        raw = raw.rename(columns={"date": "Date", "hour": "Hour",
                                  "pp_id": "EPIAS Plant ID",
                                  "production_mwh": "Production (MWh)"})
        raw[["Date", "Hour", "Plant Name", "EPIAS Plant ID", "Production (MWh)"]].head(50_000).to_excel(
            writer, sheet_name="Raw Hourly (sample)", index=False
        )

        # ── Tab 6: Metadata ───────────────────────────────────────────────
        pd.DataFrame([
            ("Data Source",    "EPİAŞ Transparency Platform — seffaflik.epias.com.tr"),
            ("API Endpoint",   "rt-gen-bulk (Plant-Level Real-Time Generation)"),
            ("Coverage Start", "May 2019 (EPİAŞ began publishing plant-level data on 16 May 2019)"),
            ("Coverage End",   (date.today() - timedelta(1)).strftime("%d %B %Y")),
            ("Granularity",    "Monthly totals (MWh) — summed from hourly EPIAS data"),
            ("Company",        "Zorlu Enerji Elektrik Uretim A.S. (ZOREN) and subsidiaries"),
            ("Date Pulled",    pull_date),
            ("Note 1",         "2019 is a partial year — data starts 16 May 2019"),
            ("Note 2",         "Gokcedag Wind (135 MW) sold to Ronesans Enerji Dec 2025 — rows after 2025-12-31 are excluded from this dataset"),
            ("Note 3",         "Pakistan (Jhimpir 56 MW) and Palestine (Dead Sea 1.5 MW) are NOT in EPIAS — not included"),
            ("Note 4",         "Mercan hydro complex is split into 3 registered units in EPIAS: Mercan (20.4 MW known), Mercan (Yukari) and Mercan (Haci) — individual capacity for sub-units not available in public sources; sum all three for total Mercan output"),
            ("Note 5",         "Lüleburgaz natural gas plant (49.5 MW) produced only 7.5 GWh total, almost entirely in 2019. Effectively inactive since — consistent with Zorlu's stated exit from fossil fuels"),
            ("Note 6",         "Solar hybrid units at Alasehir (3.75 MW) and Kizildere (0.99 MW) are not registered as separate plants in EPIAS and are NOT in this dataset"),
        ], columns=["Field", "Value"]).to_excel(writer, sheet_name="Metadata", index=False)

        # ── Tab 7: Plant Reference — static plant list with capacity ──────
        pd.DataFrame(ZORLU_PLANTS).rename(columns={
            "pp_id":       "EPIAS Plant ID",
            "plant_name":  "Plant Name",
            "fuel_type":   "Fuel Type",
            "capacity_mw": "Installed Capacity (MW)",
            "location":    "Location",
        }).to_excel(writer, sheet_name="Plant Reference", index=False)

        # ── Post-processing: widths, bold headers, number formats ───────
        book = writer.book
        MWH_FMT   = "#,##0"        # e.g.  128,176
        MW_FMT    = "#,##0.00"     # e.g.      45.00
        HEADER_FILL = PatternFill("solid", fgColor="1F3864")   # dark navy

        # Columns that hold MWh production values (by name)
        MWH_COLS  = {"Production (MWh)", "Total Production (MWh)"}
        MW_COLS   = {"Installed Capacity (MW)"}

        for sheet_name in ["Guide", "Summary", "By Plant", "Wide Pivot",
                           "By Fuel Type", "Metadata", "Plant Reference",
                           "Raw Hourly (sample)"]:
            ws = book[sheet_name]

            # Auto-fit column widths
            for col in ws.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)

            # Style header row: bold white text on dark navy background
            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = HEADER_FILL
                cell.alignment = Alignment(horizontal="center")

            # Apply number formats based on header name
            header_cells = {cell.value: cell.column for cell in ws[1]}
            for col_name, col_idx in header_cells.items():
                col_letter = get_column_letter(col_idx)
                if col_name in MWH_COLS:
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.number_format = MWH_FMT
                elif col_name in MW_COLS:
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.number_format = MW_FMT

            # Wide Pivot: all plant columns (cols 4+) are MWh
            if sheet_name == "Wide Pivot":
                for col_cells in ws.iter_cols(min_col=4, min_row=2):
                    for cell in col_cells:
                        cell.number_format = MWH_FMT

        # Freeze top row on the main tabs
        for sheet_name in ["Summary", "By Plant", "Wide Pivot", "By Fuel Type"]:
            book[sheet_name].freeze_panes = "A2"

        # Guide sheet: wrap text in description columns, wider columns
        ws_guide = book["Guide"]
        for row in ws_guide.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws_guide.column_dimensions["A"].width = 22
        ws_guide.column_dimensions["B"].width = 60
        ws_guide.column_dimensions["C"].width = 55
        ws_guide.row_dimensions[1].height = 18
        for i in range(2, ws_guide.max_row + 1):
            ws_guide.row_dimensions[i].height = 45

        # Alternate row shading on Guide for readability
        ALT_FILL = PatternFill("solid", fgColor="EEF2F7")
        for row_idx, row in enumerate(ws_guide.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = ALT_FILL

    print(f"  OK: Excel saved -> {EXCEL_OUT.name}")


# ── Entry point ───────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("  Zorlu Enerji Generation Data Fetcher")
    print("  Source: EPIAS seffaflik.epias.com.tr")
    print("=" * 60)

    eptr = get_client()
    print(f"  Connected. Fetching {len(ZORLU_PLANTS)} plants from {DATA_START} to yesterday\n")

    print("[3/4] Fetching daily generation data …")
    df_raw   = build_raw_data(eptr)
    monthly  = aggregate_to_monthly(df_raw)
    export_excel(monthly)

    # Quick sanity check
    total_gwh = monthly["production_mwh"].sum() / 1000
    n_months  = monthly[["year", "month"]].drop_duplicates().shape[0]
    print(f"\n  Plants: {monthly['plant_name'].nunique()}")
    print(f"  Months: {n_months}")
    print(f"  Total:  {total_gwh:,.0f} GWh across all plants and periods")
    print("\nDONE! Open zorlu_enerji_generation.xlsx")


if __name__ == "__main__":
    main()
