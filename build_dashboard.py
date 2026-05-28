"""
Builds the Dashboard sheet in zorlu_enerji_generation.xlsx.
Run separately: python build_dashboard.py
"""
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo
from fetch_zorlu import aggregate_to_monthly, RAW_CSV, ZORLU_PLANTS

df_raw  = pd.read_csv(RAW_CSV)
monthly = aggregate_to_monthly(df_raw)
raw_df  = df_raw   # alias used for coverage dates

annual = monthly.groupby(
    ["plant_name", "fuel_type", "capacity_mw", "year"]
)["production_mwh"].sum().reset_index()
annual["gwh"] = (annual["production_mwh"] / 1000).round(1)

pivot = annual.pivot_table(
    index=["plant_name", "fuel_type", "capacity_mw"],
    columns="year", values="gwh", fill_value=0
).reset_index()
pivot.columns = [str(c) if isinstance(c, int) else c for c in pivot.columns]
ALL_YRS = [str(y) for y in range(2019, 2027)]
pivot["Total"] = pivot[[y for y in ALL_YRS if y in pivot.columns]].sum(axis=1).round(1)
pivot = pivot.sort_values("2024", ascending=False).reset_index(drop=True)

seas = (
    monthly.groupby(["plant_name", "month"])["production_mwh"]
    .mean().round(0).unstack(fill_value=0)
)
seas.columns = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
seas = seas.set_index(seas.reset_index()["plant_name"]).reindex(pivot["plant_name"])
seas = seas.reset_index()

# Years with complete 12-month data
FULL_YEARS  = sorted([y for y in monthly["year"].unique()
                      if monthly[monthly["year"] == y]["month"].nunique() == 12])
LATEST_FULL = FULL_YEARS[-1]
AVG_LABEL   = f"Avg {FULL_YEARS[0]}-{str(LATEST_FULL)[-2:]}"
PARTIAL_YRS = {"2019", "2026"}

# Metrics derived from data (no hardcoding)
total_cap    = sum(p["capacity_mw"] for p in ZORLU_PLANTS if p["capacity_mw"])
total_rows   = len(raw_df)
dates        = pd.to_datetime(raw_df["date"])
coverage_str = f"{dates.min().strftime('%b %Y')} - {dates.max().strftime('%b %Y')}"

latest_gwh   = monthly[monthly["year"] == LATEST_FULL]["production_mwh"].sum() / 1000
top_plant    = (monthly[monthly["year"] == LATEST_FULL]
                .groupby("plant_name")["production_mwh"].sum().idxmax())
top_plant_gwh = (monthly[(monthly["year"] == LATEST_FULL) &
                          (monthly["plant_name"] == top_plant)]["production_mwh"].sum() / 1000)
top_plant_cap = next((p["capacity_mw"] for p in ZORLU_PLANTS if p["plant_name"] == top_plant), 0)
top_cf        = round(top_plant_gwh * 1000 / (top_plant_cap * 8760) * 100) if top_plant_cap else 0
top_fuel      = (monthly[monthly["year"] == LATEST_FULL]
                 .groupby("fuel_type")["production_mwh"].sum().idxmax())
top_fuel_pct  = round(
    monthly[(monthly["year"] == LATEST_FULL) & (monthly["fuel_type"] == top_fuel)]
    ["production_mwh"].sum() /
    monthly[monthly["year"] == LATEST_FULL]["production_mwh"].sum() * 100
)

NAVY   = "1F3864"
TEAL   = "2E75B6"
LGREY  = "F2F2F2"
WHITE  = "FFFFFF"
DGREEN = "375623"
MWH_FMT = "#,##0"
GWH_FMT = "#,##0.0"
MW_FMT  = "#,##0.0"
PCT_FMT = '0"%"'
NCOLS   = 18


def cell(ws, r, col, value="", bold=False, size=11, color="000000",
         bg=None, align="left", italic=False, num_fmt=None):
    cl = ws.cell(row=r, column=col, value=value)
    cl.font      = Font(bold=bold, size=size, color=color, italic=italic)
    cl.alignment = Alignment(horizontal=align, vertical="center")
    if bg:
        cl.fill = PatternFill("solid", fgColor=bg)
    if num_fmt:
        cl.number_format = num_fmt
    return cl


def cf_pct(row, year):
    cap = row["capacity_mw"]
    gwh = row.get(year, 0)
    if isinstance(gwh, pd.Series):
        gwh = gwh.iloc[0]
    return round(gwh * 1000 / (cap * 8760) * 100) if cap else None


def merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


wb = openpyxl.load_workbook("zorlu_enerji_generation.xlsx")
if "Dashboard" in wb.sheetnames:
    del wb["Dashboard"]
ws = wb.create_sheet("Dashboard", 0)
ws.sheet_view.showGridLines = False

# ── Title ─────────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 34
merge(ws, 1, 1, 1, NCOLS)
cell(ws, 1, 1, "ZORLU ENERJI -- Plant-Level Electricity Generation Dashboard",
     bold=True, size=16, color=WHITE, bg=NAVY, align="center")

ws.row_dimensions[2].height = 18
merge(ws, 2, 1, 2, NCOLS)
cell(ws, 2, 1,
     f"Source: EPIAS Transparency Platform  |  Coverage: {coverage_str}  |  Units: GWh unless noted",
     size=10, color=WHITE, bg=TEAL, align="center", italic=True)

# ── Key metrics ───────────────────────────────────────────────────────────────
for r, h in [(3, 10), (4, 20), (5, 16), (6, 38), (7, 16), (8, 10)]:
    ws.row_dimensions[r].height = h

merge(ws, 4, 1, 4, NCOLS)
cell(ws, 4, 1, f"KEY METRICS -- {LATEST_FULL}",
     bold=True, size=12, color=WHITE, bg=NAVY, align="center")

metrics = [
    ("Total Installed Capacity",      f"{total_cap:.0f} MW",         "All Turkish plants combined"),
    (f"{LATEST_FULL} Total Generation", f"{latest_gwh:,.0f} GWh",   f"Across {monthly['plant_name'].nunique()} plants"),
    (f"Largest Plant ({LATEST_FULL})", top_plant,                    f"{top_plant_gwh:,.0f} GWh  |  {top_plant_cap:.0f} MW  |  CF {top_cf}%"),
    (f"Top Fuel ({LATEST_FULL})",     top_fuel,                      f"{top_fuel_pct}% of total output"),
    ("Historical Range",              coverage_str,                  "Plant-level data from EPIAS"),
    ("Hourly Data Points",            f"{total_rows:,}",             "Aggregated to monthly MWh"),
]
for i, (label, value, note) in enumerate(metrics):
    col = 1 + i * 3
    merge(ws, 5, col, 5, col + 2)
    merge(ws, 6, col, 6, col + 2)
    merge(ws, 7, col, 7, col + 2)
    cell(ws, 5, col, label, bold=True, size=9,  color="AAAAAA", align="center")
    cell(ws, 6, col, value, bold=True, size=17, color=NAVY,     bg=LGREY, align="center")
    cell(ws, 7, col, note,  italic=True, size=8, color="777777", align="center")

# ── Annual production heatmap ─────────────────────────────────────────────────
R = 9
ws.row_dimensions[R].height = 10;  R += 1
ws.row_dimensions[R].height = 22
merge(ws, R, 1, R, NCOLS)
cell(ws, R, 1, "ANNUAL PRODUCTION BY PLANT (GWh)   * partial year",
     bold=True, size=12, color=WHITE, bg=NAVY)
R += 1

hdr_labels = (["Plant", "Fuel Type", "MW"] +
              [f"{y}*" if y in PARTIAL_YRS else y for y in ALL_YRS] + ["TOTAL"])
for ci, h in enumerate(hdr_labels, start=1):
    cell(ws, R, ci, h, bold=True, size=10, color=WHITE, bg=TEAL, align="center")
ws.row_dimensions[R].height = 18;  R += 1

DATA_START = R
for idx, row in pivot.iterrows():
    bg_row = LGREY if idx % 2 == 0 else WHITE
    ws.row_dimensions[R].height = 16
    cell(ws, R, 1, row["plant_name"], bold=True, size=10, bg=bg_row)
    cell(ws, R, 2, row["fuel_type"],  size=10, bg=bg_row)
    cap = row["capacity_mw"]
    cell(ws, R, 3, cap if cap else "", size=10, align="right", bg=bg_row, num_fmt=MW_FMT)
    for ci, yr in enumerate(ALL_YRS, start=4):
        val = row.get(yr, 0)
        cell(ws, R, ci, val if val > 0 else "", size=10, align="right", bg=bg_row, num_fmt=GWH_FMT)
    cell(ws, R, 12, row["Total"], bold=True, size=10, align="right", bg=bg_row, num_fmt=GWH_FMT)
    R += 1
DATA_END = R - 1

ws.row_dimensions[R].height = 18
cell(ws, R, 1, "TOTAL", bold=True, size=10, color=WHITE, bg=NAVY)
cell(ws, R, 2, "", bg=NAVY)
cell(ws, R, 3, f"{total_cap:.0f}", bold=True, size=10, color=WHITE, bg=NAVY, align="right")
for ci, yr in enumerate(ALL_YRS, start=4):
    tot = pivot[yr].sum() if yr in pivot.columns else 0
    cell(ws, R, ci, round(tot, 1) if tot > 0 else "",
         bold=True, size=10, color=WHITE, bg=NAVY, align="right", num_fmt=GWH_FMT)
cell(ws, R, 12, round(pivot["Total"].sum(), 1),
     bold=True, size=10, color=WHITE, bg=NAVY, align="right", num_fmt=GWH_FMT)

full_yr_cols = [4 + ALL_YRS.index(str(y)) for y in FULL_YEARS]
for col_idx in full_yr_cols:
    rng = f"{get_column_letter(col_idx)}{DATA_START}:{get_column_letter(col_idx)}{DATA_END}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=0,     start_color="FFFFFF",
        mid_type="percentile", mid_value=50, mid_color="C6EFCE",
        end_type="max",                      end_color=DGREEN))
R += 2

# ── Capacity factors ──────────────────────────────────────────────────────────
ws.row_dimensions[R].height = 22
merge(ws, R, 1, R, NCOLS)
cell(ws, R, 1, "CAPACITY FACTOR (%)   =   actual output / theoretical maximum at rated capacity",
     bold=True, size=12, color=WHITE, bg=NAVY)
R += 1

cf_yr_strs = [str(y) for y in FULL_YEARS]
avg_col    = 4 + len(FULL_YEARS)
for ci, h in enumerate(["Plant", "Fuel Type", "MW"] + cf_yr_strs + [AVG_LABEL], start=1):
    cell(ws, R, ci, h, bold=True, size=10, color=WHITE, bg=TEAL, align="center")
ws.row_dimensions[R].height = 18;  R += 1

CF_START = R
for idx, row in pivot.iterrows():
    bg_row = LGREY if idx % 2 == 0 else WHITE
    ws.row_dimensions[R].height = 16
    cap = row["capacity_mw"]
    cell(ws, R, 1, row["plant_name"], bold=True, size=10, bg=bg_row)
    cell(ws, R, 2, row["fuel_type"],  size=10, bg=bg_row)
    cell(ws, R, 3, cap if cap else "", size=10, align="right", bg=bg_row, num_fmt=MW_FMT)
    cfs = []
    for ci, yr in enumerate(cf_yr_strs, start=4):
        v = cf_pct(row, yr)
        cell(ws, R, ci, v if v is not None else "", size=10, align="right",
             bg=bg_row, num_fmt=PCT_FMT)
        if v is not None:
            cfs.append(v)
    avg = round(sum(cfs) / len(cfs)) if cfs else None
    cell(ws, R, avg_col, avg if avg is not None else "", bold=True, size=10,
         align="right", bg=bg_row, num_fmt=PCT_FMT)
    R += 1
CF_END = R - 1

for col_idx in range(4, avg_col + 1):
    rng = f"{get_column_letter(col_idx)}{CF_START}:{get_column_letter(col_idx)}{CF_END}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="num", start_value=0,  start_color="FFCCCC",
        mid_type="num",   mid_value=40,   mid_color="FFEB9C",
        end_type="num",   end_value=85,   end_color=DGREEN))
R += 2

# ── Monthly seasonality ───────────────────────────────────────────────────────
ws.row_dimensions[R].height = 22
merge(ws, R, 1, R, NCOLS)
cell(ws, R, 1,
     "AVERAGE MONTHLY PRODUCTION (MWh)   --   seasonal pattern, avg across all available years",
     bold=True, size=12, color=WHITE, bg=NAVY)
R += 1

MONTHS = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
for ci, h in enumerate(["Plant","Fuel Type"] + MONTHS + ["Annual avg GWh"], start=1):
    cell(ws, R, ci, h, bold=True, size=10, color=WHITE, bg=TEAL, align="center")
ws.row_dimensions[R].height = 18;  R += 1

SEAS_START = R
for idx, srow in seas.iterrows():
    bg_row = LGREY if idx % 2 == 0 else WHITE
    ws.row_dimensions[R].height = 16
    pname = srow["plant_name"]
    fuel  = pivot.loc[pivot["plant_name"] == pname, "fuel_type"].values
    fuel  = fuel[0] if len(fuel) else ""
    cell(ws, R, 1, pname, bold=True, size=10, bg=bg_row)
    cell(ws, R, 2, fuel,  size=10, bg=bg_row)
    total_monthly = 0
    for ci, mon in enumerate(MONTHS, start=3):
        v = srow.get(mon, 0)
        cell(ws, R, ci, int(v) if v > 0 else "", size=10, align="right",
             bg=bg_row, num_fmt=MWH_FMT)
        total_monthly += v
    cell(ws, R, 15, round(total_monthly / 1000, 1), bold=True, size=10,
         align="right", bg=bg_row, num_fmt=GWH_FMT)
    R += 1
SEAS_END = R - 1

for col_idx in range(3, 15):
    rng = f"{get_column_letter(col_idx)}{SEAS_START}:{get_column_letter(col_idx)}{SEAS_END}"
    ws.conditional_formatting.add(rng, ColorScaleRule(
        start_type="min", start_color="FFFFFF",
        end_type="max",   end_color="2E75B6"))

ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 7
for i in range(4, 19):
    ws.column_dimensions[get_column_letter(i)].width = 9

# ── Number formats on all other sheets ───────────────────────────────────────
MWH_COLS = {"Production (MWh)", "Total Production (MWh)"}
MW_COLS  = {"Installed Capacity (MW)"}

for sheet_name in wb.sheetnames:
    if sheet_name == "Dashboard":
        continue
    ws2 = wb[sheet_name]
    header_map = {ws2.cell(1, c).value: c for c in range(1, ws2.max_column + 1)}
    for col_name, col_idx in header_map.items():
        if col_name in MWH_COLS:
            for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cl in row:
                    cl.number_format = MWH_FMT
        elif col_name in MW_COLS:
            for row in ws2.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cl in row:
                    cl.number_format = MW_FMT
    if sheet_name == "Wide Pivot":
        for col_cells in ws2.iter_cols(min_col=2, min_row=2):
            for cl in col_cells:
                if isinstance(cl.value, (int, float)):
                    cl.number_format = GWH_FMT

# ── Autofilter tables on key sheets ──────────────────────────────────────────
def make_table(ws, name, ref, style="TableStyleMedium2"):
    if name in ws.tables:
        del ws.tables[name]
    tab = Table(displayName=name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style, showFirstColumn=False, showLastColumn=False,
        showRowStripes=True, showColumnStripes=False
    )
    ws.add_table(tab)

ws_bp  = wb["By Plant"]
ws_ft  = wb["By Fuel Type"]
ws_sum = wb["Summary"]
make_table(ws_bp,  "ByPlant",    f"A1:{get_column_letter(ws_bp.max_column)}{ws_bp.max_row}")
make_table(ws_ft,  "ByFuelType", f"A1:{get_column_letter(ws_ft.max_column)}{ws_ft.max_row}")
make_table(ws_sum, "Summary",    f"A1:B{ws_sum.max_row}", style="TableStyleMedium9")

if "Charts" in wb.sheetnames:
    del wb["Charts"]

wb.save("zorlu_enerji_generation.xlsx")
print("Dashboard built. Autofilter tables added to By Plant, By Fuel Type, Summary.")
